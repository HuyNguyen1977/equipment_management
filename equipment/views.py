from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.db.models import Q, Count, Exists, OuterRef
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import Company, Equipment, EquipmentHistory
from .forms import EquipmentForm, EquipmentHistoryForm


@login_required
def index(request):
    """Trang chủ - danh sách thiết bị"""
    from django.db.models import Exists, OuterRef
    
    companies = Company.objects.all()
    equipment_list = Equipment.objects.select_related('company', 'current_user').all()
    
    # Filter theo công ty
    company_id = request.GET.get('company')
    if company_id:
        equipment_list = equipment_list.filter(company_id=company_id)
    
    # Filter theo loại thiết bị
    equipment_type = request.GET.get('equipment_type')
    if equipment_type:
        equipment_list = equipment_list.filter(equipment_type=equipment_type)
    
    # Search
    search = request.GET.get('search', '').strip()
    if search:
        equipment_list = equipment_list.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(machine_name__icontains=search) |
            Q(current_user__username__icontains=search) |
            Q(current_user__first_name__icontains=search) |
            Q(current_user__last_name__icontains=search) |
            Q(company__name__icontains=search)
        )
    
    # Kiểm tra xem thiết bị nào có lịch sử thanh lý (tối ưu query)
    from equipment.models import EquipmentHistory
    liquidation_exists = EquipmentHistory.objects.filter(
        equipment=OuterRef('pk'),
        action_type='liquidation'
    )
    equipment_list = equipment_list.annotate(has_liquidation=Exists(liquidation_exists))
    
    # Phân trang
    paginator = Paginator(equipment_list, 20)  # 20 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'companies': companies,
        'equipment_list': page_obj,
        'page_obj': page_obj,
        'selected_company': company_id,
        'selected_type': equipment_type,
        'search_query': search,
    }
    return render(request, 'equipment/index.html', context)


@login_required
def equipment_detail(request, pk):
    """Chi tiết thiết bị"""
    equipment = get_object_or_404(Equipment, pk=pk)
    histories = equipment.histories.all()
    
    # Lấy danh sách người đã sử dụng (từ history)
    user_histories = histories.filter(
        action_type__in=['user_assignment', 'user_return']
    ).order_by('-action_date')
    
    # Kiểm tra xem có lịch sử thanh lý không
    has_liquidation = histories.filter(action_type='liquidation').exists()
    
    # Đếm số người đã sử dụng (unique users)
    users_used = set()
    for history in user_histories:
        # Parse description để lấy tên user
        desc = history.description
        if 'Giao máy cho' in desc or 'Chuyển máy' in desc:
            # Extract username từ description
            if 'sang' in desc:
                # Chuyển máy từ A sang B
                parts = desc.split('sang')
                if len(parts) > 1:
                    users_used.add(parts[1].strip())
            elif 'Giao máy cho' in desc:
                parts = desc.split('Giao máy cho')
                if len(parts) > 1:
                    users_used.add(parts[1].strip())
        elif 'Trả máy từ' in desc:
            parts = desc.split('Trả máy từ')
            if len(parts) > 1:
                users_used.add(parts[1].strip())
    
    context = {
        'equipment': equipment,
        'histories': histories,
        'user_histories': user_histories,
        'users_count': len(users_used),
        'has_liquidation': has_liquidation,
    }
    return render(request, 'equipment/equipment_detail.html', context)


def login_view(request):
    """Đăng nhập"""
    if request.user.is_authenticated:
        return redirect('equipment:index')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Đăng nhập thành công! Chào mừng {user.username}!')
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('equipment:index')
    else:
        form = AuthenticationForm()
    
    return render(request, 'equipment/login.html', {'form': form})


def logout_view(request):
    """Đăng xuất"""
    logout(request)
    messages.success(request, 'Đã đăng xuất thành công!')
    return redirect('equipment:index')


@login_required
def search_users(request):
    """AJAX endpoint để tìm kiếm users"""
    query = request.GET.get('q', '').strip()
    if not query:
        return JsonResponse({'users': []})
    
    users = User.objects.filter(
        Q(username__icontains=query) |
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(email__icontains=query)
    ).filter(is_active=True)[:10]
    
    results = []
    for user in users:
        display_name = user.get_full_name() or user.username
        results.append({
            'id': user.id,
            'username': user.username,
            'display': f"{user.username} ({display_name})",
            'full_name': display_name
        })
    
    return JsonResponse({'users': results})


@login_required
@require_http_methods(["POST"])
def create_user_quick(request):
    """Tạo user nhanh - chỉ cần Name và Email, có thể thêm Username"""
    name = request.POST.get('name', '').strip()
    email = request.POST.get('email', '').strip()
    username = request.POST.get('username', '').strip()
    
    if not name:
        return JsonResponse({'success': False, 'error': 'Tên là bắt buộc'}, status=400)
    
    if not email:
        return JsonResponse({'success': False, 'error': 'Email là bắt buộc'}, status=400)
    
    # Nếu có username từ form, validate và sử dụng
    if username:
        import re
        # Validate username chỉ chứa chữ cái, số và dấu gạch dưới
        if not re.match(r'^[a-zA-Z0-9_]+$', username):
            return JsonResponse({'success': False, 'error': 'Tên đăng nhập chỉ được chứa chữ cái, số và dấu gạch dưới (_)'}, status=400)
        
        # Kiểm tra username đã tồn tại chưa
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': f'Tên đăng nhập "{username}" đã tồn tại. Vui lòng chọn tên khác.'}, status=400)
    else:
        # Tạo username từ name (loại bỏ dấu, khoảng trắng, chuyển thành chữ thường)
        import re
        username = re.sub(r'[^a-zA-Z0-9]', '', name.lower())
        if not username:
            username = 'user' + str(User.objects.count() + 1)
        
        # Đảm bảo username là duy nhất
        original_username = username
        counter = 1
        while User.objects.filter(username=username).exists():
            username = f"{original_username}{counter}"
            counter += 1
    
    try:
        # Tách name thành first_name và last_name
        name_parts = name.split(maxsplit=1)
        first_name = name_parts[0] if name_parts else name
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        
        # Tạo user mới với password mặc định Pega@2025
        user = User.objects.create_user(
            username=username,
            email=email,
            first_name=first_name,
            last_name=last_name,
            password='Pega@2025'  # Password mặc định
        )
        
        display_name = user.get_full_name() or user.username
        return JsonResponse({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'display': f"{user.username} ({display_name})",
                'full_name': display_name
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


def get_next_code(request):
    """API để lấy mã tiếp theo dựa trên miền - không cần login để preview code"""
    region = request.GET.get('region', '').strip()
    
    if not region:
        return JsonResponse({'success': False, 'error': 'Region is required'}, status=400)
    
    if region not in ['MN', 'MT', 'MB']:
        return JsonResponse({'success': False, 'error': 'Invalid region'}, status=400)
    
    import re
    prefix = region
    
    # Tìm số lớn nhất trong miền này
    existing_codes = Equipment.objects.filter(
        code__startswith=f"{prefix}-"
    ).values_list('code', flat=True)
    
    max_num = 0
    for code in existing_codes:
        match = re.match(rf'^{prefix}-(\d+)$', code)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
    
    # Tạo code mới
    new_num = max_num + 1
    next_code = f"{prefix}-{new_num:03d}"
    
    return JsonResponse({
        'success': True,
        'code': next_code,
        'region': region
    })


@login_required
@staff_member_required
def equipment_create(request):
    """Tạo mới thiết bị"""
    if request.method == 'POST':
        form = EquipmentForm(request.POST, request.FILES)
        if form.is_valid():
            equipment = form.save()
            
            # Tạo history nếu có người dùng được gán
            if equipment.current_user:
                EquipmentHistory.objects.create(
                    equipment=equipment,
                    action_date=date.today(),
                    action_type='user_assignment',
                    description=f'Giao máy cho {equipment.current_user.get_full_name() or equipment.current_user.username}',
                    signed_by=request.user.get_full_name() or request.user.username
                )
            
            messages.success(request, f'Đã tạo thiết bị {equipment.name} thành công!')
            return redirect('equipment:equipment_detail', pk=equipment.pk)
    else:
        form = EquipmentForm()
    
    context = {
        'form': form,
        'title': 'Tạo mới thiết bị',
    }
    return render(request, 'equipment/equipment_form.html', context)


@login_required
@staff_member_required
def equipment_edit(request, pk):
    """Chỉnh sửa thiết bị"""
    equipment = get_object_or_404(Equipment, pk=pk)
    old_user = equipment.current_user  # Lưu user cũ trước khi thay đổi
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, request.FILES, instance=equipment)
        if form.is_valid():
            equipment = form.save()
            new_user = equipment.current_user
            
            # Tạo history nếu có thay đổi người dùng
            if old_user != new_user:
                if new_user and not old_user:
                    # Giao máy mới
                    EquipmentHistory.objects.create(
                        equipment=equipment,
                        action_date=date.today(),
                        action_type='user_assignment',
                        description=f'Giao máy cho {new_user.get_full_name() or new_user.username}',
                        signed_by=request.user.get_full_name() or request.user.username
                    )
                elif old_user and not new_user:
                    # Trả máy
                    EquipmentHistory.objects.create(
                        equipment=equipment,
                        action_date=date.today(),
                        action_type='user_return',
                        description=f'Trả máy từ {old_user.get_full_name() or old_user.username}',
                        signed_by=request.user.get_full_name() or request.user.username
                    )
                elif old_user and new_user:
                    # Chuyển máy từ người này sang người khác
                    EquipmentHistory.objects.create(
                        equipment=equipment,
                        action_date=date.today(),
                        action_type='user_assignment',
                        description=f'Chuyển máy từ {old_user.get_full_name() or old_user.username} sang {new_user.get_full_name() or new_user.username}',
                        signed_by=request.user.get_full_name() or request.user.username
                    )
            
            messages.success(request, f'Đã cập nhật thiết bị {equipment.name} thành công!')
            return redirect('equipment:equipment_detail', pk=equipment.pk)
    else:
        form = EquipmentForm(instance=equipment)
    
    context = {
        'form': form,
        'equipment': equipment,
        'title': 'Chỉnh sửa thiết bị',
    }
    return render(request, 'equipment/equipment_form.html', context)


@login_required
@staff_member_required
def equipment_delete(request, pk):
    """Xóa thiết bị"""
    equipment = get_object_or_404(Equipment, pk=pk)
    
    if request.method == 'POST':
        equipment.delete()
        messages.success(request, f'Đã xóa thiết bị {equipment.name}!')
        return redirect('equipment:index')
    
    context = {
        'equipment': equipment,
    }
    return render(request, 'equipment/equipment_delete.html', context)


@login_required
def equipment_history(request, pk):
    """Xem lịch sử thiết bị theo mẫu"""
    equipment = get_object_or_404(Equipment, pk=pk)
    histories = equipment.histories.all()
    
    context = {
        'equipment': equipment,
        'histories': histories,
    }
    return render(request, 'equipment/equipment_history.html', context)


@login_required
@staff_member_required
def history_add(request, equipment_pk):
    """Thêm lịch sử cho thiết bị"""
    equipment = get_object_or_404(Equipment, pk=equipment_pk)
    action_type = request.GET.get('action_type', '')  # Lấy action_type từ URL nếu có
    
    if request.method == 'POST':
        form = EquipmentHistoryForm(request.POST)
        if form.is_valid():
            history = form.save(commit=False)
            history.equipment = equipment
            
            # Tự động cập nhật trạng thái thiết bị dựa trên loại hành động
            action_type_value = form.cleaned_data.get('action_type')
            description = form.cleaned_data.get('description', '')
            to_user = form.cleaned_data.get('to_user')
            
            if action_type_value == 'movement' and to_user:
                # Di chuyển → cập nhật người sử dụng mới
                old_user = equipment.current_user
                equipment.current_user = to_user
                equipment.is_active = True
                
                # Tự động điền description nếu chưa có hoặc rỗng
                if not description or description.strip() == '':
                    if old_user:
                        description = f'Chuyển máy từ {old_user.get_full_name() or old_user.username} sang {to_user.get_full_name() or to_user.username}'
                    else:
                        description = f'Chuyển máy sang {to_user.get_full_name() or to_user.username}'
                    history.description = description
                
                equipment.save()
            elif action_type_value == 'liquidation':
                # Thanh lý → hết hoạt động, không còn người sử dụng
                equipment.is_active = False
                equipment.current_user = None
                equipment.save()
            elif action_type_value == 'user_assignment':
                # Giao máy cho người dùng → đang hoạt động
                equipment.is_active = True
                # Cố gắng lấy user từ description
                # Format: "Giao máy cho username" hoặc "Chuyển máy từ A sang B"
                from django.contrib.auth.models import User
                if 'Giao máy cho' in description:
                    parts = description.split('Giao máy cho')
                    if len(parts) > 1:
                        username = parts[1].strip()
                        try:
                            user = User.objects.get(username=username)
                            equipment.current_user = user
                        except User.DoesNotExist:
                            pass
                elif 'sang' in description:
                    parts = description.split('sang')
                    if len(parts) > 1:
                        username = parts[1].strip()
                        try:
                            user = User.objects.get(username=username)
                            equipment.current_user = user
                        except User.DoesNotExist:
                            pass
                equipment.save()
            elif action_type_value == 'user_return':
                # Trả máy → không còn người sử dụng (nhưng có thể vẫn hoạt động)
                equipment.current_user = None
                equipment.save()
            
            history.save()
            messages.success(request, 'Đã thêm lịch sử thiết bị thành công!')
            return redirect('equipment:equipment_history', pk=equipment.pk)
    else:
        form = EquipmentHistoryForm()
        # Nếu có action_type trong URL, set giá trị mặc định
        if action_type:
            form.fields['action_type'].initial = action_type
    
    context = {
        'form': form,
        'equipment': equipment,
        'action_type': action_type,
    }
    return render(request, 'equipment/history_form.html', context)


@login_required
@staff_member_required
def history_edit(request, pk):
    """Chỉnh sửa lịch sử thiết bị"""
    history = get_object_or_404(EquipmentHistory, pk=pk)
    old_action_type = history.action_type  # Lưu action_type cũ
    equipment = history.equipment
    
    if request.method == 'POST':
        form = EquipmentHistoryForm(request.POST, instance=history)
        if form.is_valid():
            # Tự động cập nhật trạng thái thiết bị dựa trên loại hành động
            action_type_value = form.cleaned_data.get('action_type')
            description = form.cleaned_data.get('description', '')
            to_user = form.cleaned_data.get('to_user')
            
            if action_type_value == 'movement' and to_user:
                # Di chuyển → cập nhật người sử dụng mới
                old_user = equipment.current_user
                equipment.current_user = to_user
                equipment.is_active = True
                
                # Tự động điền description nếu chưa có hoặc rỗng
                if not description or description.strip() == '':
                    if old_user:
                        description = f'Chuyển máy từ {old_user.get_full_name() or old_user.username} sang {to_user.get_full_name() or to_user.username}'
                    else:
                        description = f'Chuyển máy sang {to_user.get_full_name() or to_user.username}'
                    # Cập nhật description trong instance trước khi save
                    history.description = description
                
                equipment.save()
            elif action_type_value == 'liquidation':
                # Thanh lý → hết hoạt động, không còn người sử dụng
                equipment.is_active = False
                equipment.current_user = None
                equipment.save()
            elif action_type_value == 'user_assignment':
                # Giao máy cho người dùng → đang hoạt động
                equipment.is_active = True
                # Cố gắng lấy user từ description
                from django.contrib.auth.models import User
                if 'Giao máy cho' in description:
                    parts = description.split('Giao máy cho')
                    if len(parts) > 1:
                        username = parts[1].strip()
                        try:
                            user = User.objects.get(username=username)
                            equipment.current_user = user
                        except User.DoesNotExist:
                            pass
                elif 'sang' in description:
                    parts = description.split('sang')
                    if len(parts) > 1:
                        username = parts[1].strip()
                        try:
                            user = User.objects.get(username=username)
                            equipment.current_user = user
                        except User.DoesNotExist:
                            pass
                equipment.save()
            elif action_type_value == 'user_return':
                # Trả máy → không còn người sử dụng
                equipment.current_user = None
                equipment.save()
            
            # Lưu history (với commit=False nếu cần cập nhật description)
            if action_type_value == 'movement' and to_user and (not description or description.strip() == ''):
                history = form.save(commit=False)
                # Description đã được cập nhật ở trên
                history.save()
            else:
                history = form.save()
            
            messages.success(request, 'Đã cập nhật lịch sử thiết bị thành công!')
            return redirect('equipment:equipment_history', pk=equipment.pk)
    else:
        form = EquipmentHistoryForm(instance=history)
    
    context = {
        'form': form,
        'history': history,
        'equipment': equipment,
    }
    return render(request, 'equipment/history_form.html', context)


@login_required
@staff_member_required
def history_delete(request, pk):
    """Xóa lịch sử thiết bị"""
    history = get_object_or_404(EquipmentHistory, pk=pk)
    equipment_pk = history.equipment.pk
    
    if request.method == 'POST':
        history.delete()
        messages.success(request, 'Đã xóa lịch sử thiết bị!')
        return redirect('equipment_history', pk=equipment_pk)
    
    context = {
        'history': history,
    }
    return render(request, 'equipment/history_delete.html', context)


@login_required
def report(request):
    """Trang báo cáo thiết bị với khả năng xuất Excel"""
    from django.db.models import Exists, OuterRef
    
    # Lấy tất cả thiết bị
    equipment_list = Equipment.objects.select_related('company', 'current_user').all()
    
    # Filter theo công ty
    company_id = request.GET.get('company')
    if company_id:
        equipment_list = equipment_list.filter(company_id=company_id)
    
    # Filter theo loại thiết bị
    equipment_type = request.GET.get('equipment_type')
    if equipment_type:
        equipment_list = equipment_list.filter(equipment_type=equipment_type)
    
    # Filter theo miền
    region = request.GET.get('region')
    if region:
        equipment_list = equipment_list.filter(region=region)
    
    # Filter theo trạng thái
    status = request.GET.get('status')
    if status == 'active':
        equipment_list = equipment_list.filter(is_active=True)
    elif status == 'inactive':
        equipment_list = equipment_list.filter(is_active=False)
    
    # Search
    search = request.GET.get('search', '').strip()
    if search:
        equipment_list = equipment_list.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(machine_name__icontains=search) |
            Q(current_user__username__icontains=search) |
            Q(current_user__first_name__icontains=search) |
            Q(current_user__last_name__icontains=search) |
            Q(company__name__icontains=search)
        )
    
    # Kiểm tra xem thiết bị nào có lịch sử thanh lý
    liquidation_exists = EquipmentHistory.objects.filter(
        equipment=OuterRef('pk'),
        action_type='liquidation'
    )
    equipment_list = equipment_list.annotate(has_liquidation=Exists(liquidation_exists))
    
    # Nếu có parameter export=excel, xuất file Excel
    if request.GET.get('export') == 'excel':
        return export_to_excel(equipment_list)
    
    # Thống kê tổng quan
    total_equipment = Equipment.objects.count()
    active_equipment = Equipment.objects.filter(is_active=True).count()
    inactive_equipment = Equipment.objects.filter(is_active=False).count()
    
    # Đếm theo miền
    region_stats = Equipment.objects.values('region').annotate(count=Count('id'))
    region_dict = {r['region']: r['count'] for r in region_stats}
    
    # Đếm theo loại thiết bị
    type_stats = Equipment.objects.values('equipment_type').annotate(count=Count('id'))
    type_dict = {t['equipment_type']: t['count'] for t in type_stats}
    
    # Đếm thiết bị đang được sử dụng
    in_use = Equipment.objects.filter(current_user__isnull=False).count()
    
    # Đếm thiết bị đã thanh lý
    liquidation_equipment = Equipment.objects.filter(
        histories__action_type='liquidation'
    ).distinct().count()
    
    companies = Company.objects.all()
    
    context = {
        'equipment_list': equipment_list,
        'companies': companies,
        'selected_company': company_id,
        'selected_type': equipment_type,
        'selected_region': region,
        'selected_status': status,
        'search_query': search,
        'total_equipment': total_equipment,
        'active_equipment': active_equipment,
        'inactive_equipment': inactive_equipment,
        'region_stats': region_dict,
        'type_stats': type_dict,
        'in_use': in_use,
        'liquidation_equipment': liquidation_equipment,
    }
    return render(request, 'equipment/report.html', context)


def export_to_excel(equipment_list):
    """Xuất danh sách thiết bị ra file Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo thiết bị"
    
    # Định nghĩa style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Tiêu đề
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'BÁO CÁO THIẾT BỊ IT'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Header row
    headers = [
        'STT', 'Mã thiết bị', 'Tên thiết bị', 'Công ty', 'Miền', 'Loại thiết bị',
        'Tên máy', 'Hệ điều hành', 'Nhà sản xuất', 'Model', 'Bộ xử lý',
        'Bộ nhớ', 'HDD/SSD', 'Card đồ họa', 'Người sử dụng', 'Email', 'Trạng thái'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_style
        cell.alignment = center_alignment
    
    # Dữ liệu
    row_num = 4
    for idx, equipment in enumerate(equipment_list, 1):
        # Kiểm tra trạng thái
        has_liquidation = getattr(equipment, 'has_liquidation', False)
        if has_liquidation:
            status = 'Đã thanh lý'
        elif equipment.is_active:
            status = 'Hoạt động'
        else:
            status = 'Không hoạt động'
        
        # Người sử dụng
        current_user = ''
        user_email = ''
        if equipment.current_user:
            current_user = equipment.current_user.get_full_name() or equipment.current_user.username
            user_email = equipment.current_user.email or ''
        
        # Miền
        region_display = dict(Equipment.REGIONS).get(equipment.region, equipment.region)
        
        # Loại thiết bị
        type_display = dict(Equipment.EQUIPMENT_TYPES).get(equipment.equipment_type, equipment.equipment_type)
        
        data = [
            idx,
            equipment.code,
            equipment.name,
            equipment.company.name if equipment.company else '',
            region_display,
            type_display,
            equipment.machine_name or '',
            equipment.operating_system or '',
            equipment.system_manufacturer or '',
            equipment.system_model or '',
            equipment.processor or '',
            equipment.memory or '',
            equipment.storage or '',
            equipment.graphics_card or '',
            current_user,
            user_email,
            status,
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_style
            if col_num == 1:  # STT
                cell.alignment = center_alignment
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        row_num += 1
    
    # Điều chỉnh độ rộng cột
    column_widths = [6, 15, 25, 20, 12, 15, 20, 20, 20, 20, 25, 15, 20, 20, 20, 25, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Đặt chiều cao hàng
    for row in range(3, row_num):
        ws.row_dimensions[row].height = 25
    
    # Tạo response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'bao_cao_thiet_bi_{date.today().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

